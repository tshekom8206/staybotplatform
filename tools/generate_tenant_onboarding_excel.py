import os
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
try:
    import psycopg2
    import psycopg2.extras
except Exception:
    psycopg2 = None

# Helper constants
HEADER_FILL = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")
HINT_FONT = Font(color="FF666666", italic=True)
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)

MAX_ROWS = 1000  # Room for client entries


def add_header(ws, headers, required_indices=None, hints=None):
    # Row 1: headers
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    # Row 2: hints/examples
    if hints:
        for col, hint in enumerate(hints, start=1):
            if hint is None:
                continue
            cell = ws.cell(row=2, column=col, value=hint)
            cell.font = HINT_FONT
            cell.alignment = CENTER

    # Conditional formatting for required fields (blank from row 3 downward)
    if required_indices:
        for idx in required_indices:
            col_letter = ws.cell(row=1, column=idx).column_letter
            # Highlight blanks in required columns
            ws.conditional_formatting.add(
                f"{col_letter}3:{col_letter}{MAX_ROWS}",
                CellIsRule(operator='equal', formula=['""'], stopIfTrue=False, fill=REQUIRED_FILL)
            )

    ws.freeze_panes = "A3"

def dictfetch_all(cur):
    cols = [desc[0] for desc in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]

def safe_query(conn, sql, params=None):
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params or [])
            return dictfetch_all(cur)
    except Exception:
        return []

def try_queries(conn, queries, params=None, debug=False, label=""):
    for q in queries:
        rows = safe_query(conn, q, params)
        if debug:
            print(f"[prefill] {label} -> {len(rows)} rows using: {q[:60]}...")
        if rows:
            return rows
    return []

def _parse_conn_str(conn_str: str):
    if not conn_str:
        return None
    # Detect Npgsql-style "Key=Value;" pairs and map to psycopg2
    if ";" in conn_str and "Host=" in conn_str:
        parts = [p.strip() for p in conn_str.split(";") if p.strip()]
        kv = {}
        for p in parts:
            if "=" not in p:
                continue
            k, v = p.split("=", 1)
            kv[k.strip().lower()] = v.strip()
        mapping = {
            "host": "host",
            "database": "dbname",
            "username": "user",
            "user id": "user",
            "password": "password",
            "ssl mode": "sslmode",
        }
        out = {}
        for k, v in kv.items():
            mk = mapping.get(k)
            if not mk:
                continue
            if mk == "sslmode":
                # map common values
                v = {"require": "require", "required": "require", "disable": "disable", "prefer": "prefer"}.get(v.lower(), v)
            out[mk] = v
        # Build DSN
        dsn = " ".join(f"{k}={v}" for k, v in out.items())
        return dsn
    return None

def get_connection(conn_str, debug=False):
    if not psycopg2:
        return None
    try:
        return psycopg2.connect(conn_str, cursor_factory=psycopg2.extras.DictCursor)
    except Exception as e1:
        dsn = _parse_conn_str(conn_str)
        if dsn:
            try:
                if debug:
                    print("[prefill] Retrying connection with translated DSN")
                return psycopg2.connect(dsn, cursor_factory=psycopg2.extras.DictCursor)
            except Exception as e2:
                if debug:
                    print(f"[prefill] Connection failed: {e2}")
                return None
        if debug:
            print(f"[prefill] Connection failed: {e1}")
        return None
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 28

def set_col_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        if width is None:
            continue
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = width


def add_validation_list(ws, col_idx, first_row, last_row, list_ref):
    col_letter = ws.cell(row=1, column=col_idx).column_letter
    dv = DataValidation(type="list", formula1=list_ref, allow_blank=True, showErrorMessage=True)
    dv.error = "Please select a value from the list."
    dv.prompt = "Select from dropdown"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


def build_lists_sheet(wb):
    lists = wb.create_sheet("Lists")
    lists.sheet_state = "hidden"

    data = {
        'Plans': ["Basic", "Standard", "Premium"],
        'Statuses': ["Active", "Suspended", "Cancelled"],
        'YesNo': ["TRUE", "FALSE"],
        'Departments': ["FrontDesk", "Housekeeping", "Maintenance", "Concierge", "FoodService", "Security", "IT", "Spa", "General"],
        'TaskPriority': ["Low", "Normal", "High", "Urgent"],
        'MealTypes': ["breakfast", "lunch", "dinner", "all"],
        'IntentNames': ["CANCELLATION", "CHECK_IN_OUT", "DIRECTIONS", "FEEDBACK", "RECOMMENDATION"],
        'SeverityLevels': ["Low", "Medium", "High", "Critical"],
        'TriggerCondition': ["IMMEDIATE", "SEVERITY_HIGH", "MANUAL"],
        'Currencies': ["ZAR", "USD", "EUR", "GBP"],
        'InfoCategories': ["hours", "location", "amenities", "policies", "contact"],
        'FAQCategories': ["general", "menu", "services", "amenities", "local"],
        'ServiceCategories': [
            "Local Tours", "Transportation", "Accommodation", "Business", "Dining", "Food & Beverage", "Electronics", "Amenities", "Laundry", "Concierge", "Information", "Wellness", "Spa"
        ],
    }

    named_ranges = {}
    row = 1
    for name, values in data.items():
        col = 1
        for v in values:
            lists.cell(row=row, column=col, value=v)
            col += 1
        # Define named range for this row
        start = lists.cell(row=row, column=1).coordinate
        end = lists.cell(row=row, column=len(values)).coordinate
        # Note: create_named_range is deprecated in openpyxl>=3.1, but retained here for compatibility.
        wb.create_named_range(name, lists, f"{start}:{end}")
        named_ranges[name] = f"={name}"
        row += 1

    return named_ranges

def build_start_here(wb):
    ws = wb.create_sheet("Start Here", 0)
    ws.merge_cells("A1:F1")
    ws["A1"] = "Tenant Onboarding Questionnaire — Instructions"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left")

    rows = [
        ["What is this?", "Use this workbook to provide all information needed to configure your hotel on the StayBot platform."],
        ["How to use", "Each sheet collects a topic (e.g., Core Tenant, WhatsApp). Row 1 has headers, Row 2 shows examples. Start filling from Row 3."],
        ["Required fields", "Columns marked with * are required. Blank required cells turn light red using conditional formatting."],
        ["Dropdowns", "Some columns have dropdowns. Click the cell to pick from available options."],
        ["JSON fields", "Where indicated (e.g., SupportedLanguages, Features), enter JSON like [\"en\",\"af\"]."],
        ["Minimal set", "At minimum complete: Core Tenant, WhatsApp (WABA), Owner & Staff, Hotel Info, Tenant Settings, Departments."],
        ["Need help?", "Add clarifying notes directly below the relevant row or highlight cells in yellow for review."],
        ["Sheets overview", "Core Tenant: basic account. WhatsApp: messaging setup. Owner & Staff: users. Hotel Info: property details. Tenant Settings: hours & check-in/out. Departments: teams. Others are optional and can be filled later."]
    ]

    ws.append(["Topic", "Details"])
    ws["A2"].font = HEADER_FONT
    ws["B2"].font = HEADER_FONT
    for r in rows:
        ws.append(r)

    set_col_widths(ws, [24, 120])
    ws.freeze_panes = "A3"


def sheet_core_tenant(wb, named):
    ws = wb.create_sheet("Core Tenant")
    headers = [
        "Name*", "Slug*", "Timezone", "Plan", "ThemePrimary", "Status", "RetentionDays"
    ]
    hints = [
        "StayBot Hotel Sandton", "staybot-sandton", "Africa/Johannesburg", "Basic/Standard/Premium", "#007bff", "Active", "30"
    ]
    required = [1, 2]
    add_header(ws, headers, required, hints)

    # Validations
    add_validation_list(ws, 4, 3, MAX_ROWS, named['Plans'])
    add_validation_list(ws, 6, 3, MAX_ROWS, named['Statuses'])

    # Header comments
    ws.cell(1, 1).comment = Comment("Property legal/brand name as you want it shown to guests.", "Hostr")
    ws.cell(1, 2).comment = Comment("Lowercase unique code for URLs, e.g., staybot-sandton.", "Hostr")
    ws.cell(1, 3).comment = Comment("IANA timezone. Default Africa/Johannesburg if unsure.", "Hostr")
    ws.cell(1, 4).comment = Comment("Subscription plan. If unknown, leave blank.", "Hostr")
    ws.cell(1, 5).comment = Comment("Primary brand color hex (e.g., #007bff).", "Hostr")
    ws.cell(1, 6).comment = Comment("Account status. Usually Active.", "Hostr")
    ws.cell(1, 7).comment = Comment("How long to retain guest chat data (days).", "Hostr")

    set_col_widths(ws, [34, 28, 26, 16, 16, 16, 18])

    return ws


def sheet_waba(wb, named):
    ws = wb.create_sheet("WhatsApp (WABA)")
    headers = ["WabaId*", "PhoneNumberId*", "PageAccessToken*", "VerifyToken", "Status"]
    hints = ["1234567890", "1234567890", "EAAG...", "optional", "Active"]
    required = [1, 2, 3]
    add_header(ws, headers, required, hints)
    add_validation_list(ws, 5, 3, MAX_ROWS, named['Statuses'])
    ws.cell(1, 1).comment = Comment("Facebook Business Manager WABA ID.", "Hostr")
    ws.cell(1, 2).comment = Comment("WhatsApp Phone Number ID from Meta.", "Hostr")
    ws.cell(1, 3).comment = Comment("Long-lived Page Access Token with WhatsApp permissions.", "Hostr")
    ws.cell(1, 4).comment = Comment("Used only for webhook verification (optional).", "Hostr")
    set_col_widths(ws, [20, 22, 48, 18, 14])
    return ws


def sheet_owner_staff(wb, named):
    ws = wb.create_sheet("Owner & Staff")
    headers = ["OwnerEmail*", "OwnerFirstName*", "OwnerLastName*", "OwnerPhone", "StaffEmail", "StaffFirstName", "StaffLastName", "StaffRole", "StaffPhone"]
    hints = ["owner@hotel.com", "Jane", "Doe", "+27...", "agent@hotel.com", "Sam", "Smith", "Manager/FrontDesk/...", "+27..."]
    required = [1, 2, 3]
    add_header(ws, headers, required, hints)
    ws.cell(1, 1).comment = Comment("Primary account owner email (will receive admin access).", "Hostr")
    ws.cell(1, 8).comment = Comment("Choose a staff role or specify your own if not listed.", "Hostr")
    set_col_widths(ws, [30, 18, 18, 16, 30, 18, 18, 22, 16])
    return ws


def sheet_hotel_info(wb, named):
    ws = wb.create_sheet("Hotel Info")
    headers = [
        "Description", "Category", "LogoUrl",
        "Phone", "Email", "Website",
        "Street", "City", "State", "PostalCode", "Country",
        "CheckInTime", "CheckOutTime",
        "NumberOfRooms", "NumberOfFloors", "EstablishedYear",
        "SupportedLanguages (JSON)", "DefaultLanguage", "Features (JSON)",
        "FacebookUrl", "TwitterUrl", "InstagramUrl", "LinkedInUrl",
        "CancellationPolicy", "PetPolicy", "SmokingPolicy", "ChildPolicy",
        "AllowOnlineBooking", "RequirePhoneVerification", "EnableNotifications", "EnableChatbot", "Currency"
    ]
    hints = [
        "Boutique hotel near...", "luxury/premium/...", "https://.../logo.png",
        "+27...", "info@hotel.com", "https://hotel.com",
        "123 Main St", "Johannesburg", "Gauteng", "2000", "South Africa",
        "15:00", "12:00",
        "100", "5", "2005",
        "[\"en\",\"af\"]", "en", "[\"pool\",\"spa\",\"gym\"]",
        "https://fb.com/...", "https://x.com/...", "https://instagram.com/...", "https://linkedin.com/...",
        "Free cancellation 24h", "No pets", "Non-smoking rooms", "Children welcome",
        "TRUE", "TRUE", "TRUE", "TRUE", "ZAR"
    ]
    add_header(ws, headers, [], hints)

    # Validations
    add_validation_list(ws, 31, 3, MAX_ROWS, named['Currencies'])
    add_validation_list(ws, 27, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 28, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 29, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 30, 3, MAX_ROWS, named['YesNo'])
    ws.cell(1, 1).comment = Comment("Short descriptive overview shown to guests.", "Hostr")
    ws.cell(1, 12).comment = Comment("Check-in time, 24h format HH:MM.", "Hostr")
    ws.cell(1, 13).comment = Comment("Check-out time, 24h format HH:MM.", "Hostr")
    ws.cell(1, 17).comment = Comment("JSON array of supported languages, e.g., [\"en\",\"af\"].", "Hostr")
    ws.cell(1, 19).comment = Comment("JSON array of features, e.g., [\"pool\",\"spa\"].", "Hostr")
    set_col_widths(ws, [36, 16, 26, 16, 26, 28, 26, 18, 16, 16, 18, 12, 12, 16, 18, 18, 34, 10, 34, 24, 24, 24, 24, 28, 24, 24, 24, 10, 10, 14, 10])
    return ws


def sheet_tenant_settings(wb, named):
    ws = wb.create_sheet("Tenant Settings")
    headers = [
        "BusinessHoursStart", "BusinessHoursEnd", "StandardCheckInTime", "StandardCheckOutTime",
        "LateCheckoutFeePerHour", "EarlyCheckInFeePerHour", "DefaultCurrency", "Timezone"
    ]
    hints = ["08:00", "22:00", "15:00", "12:00", "150.00", "150.00", "ZAR", "Africa/Johannesburg"]
    add_header(ws, headers, [], hints)
    add_validation_list(ws, 7, 3, MAX_ROWS, named['Currencies'])
    ws.cell(1, 1).comment = Comment("Business day start (HH:MM).", "Hostr")
    ws.cell(1, 2).comment = Comment("Business day end (HH:MM).", "Hostr")
    ws.cell(1, 5).comment = Comment("Late checkout fee per hour. Leave blank if free.", "Hostr")
    ws.cell(1, 6).comment = Comment("Early check-in fee per hour. Leave blank if free.", "Hostr")
    set_col_widths(ws, [14, 14, 18, 18, 22, 24, 12, 24])
    return ws


def sheet_departments(wb, named):
    ws = wb.create_sheet("Departments")
    headers = ["DepartmentName*", "Description", "IsActive", "Priority", "ContactInfo", "WorkingHours", "MaxConcurrentTasks"]
    hints = ["FrontDesk/Housekeeping/...", "", "TRUE", "0", "ext 100", "24/7", "10"]
    required = [1]
    add_header(ws, headers, required, hints)
    add_validation_list(ws, 1, 3, MAX_ROWS, named['Departments'])
    add_validation_list(ws, 3, 3, MAX_ROWS, named['YesNo'])
    ws.cell(1, 1).comment = Comment("Choose from common departments or enter your own.", "Hostr")
    ws.cell(1, 7).comment = Comment("Optional capacity limit of simultaneous tasks.", "Hostr")
    set_col_widths(ws, [20, 24, 10, 10, 18, 16, 18])
    return ws


def sheet_service_mapping(wb, named):
    ws = wb.create_sheet("Service-Department Mapping")
    headers = ["ServiceCategory*", "TargetDepartment*", "RequiresRoomDelivery", "RequiresAdvanceBooking", "ContactMethod", "Priority", "SpecialInstructions", "IsActive"]
    hints = ["Dining/Transportation/...", "FrontDesk/...", "FALSE", "FALSE", "phone/app/front-desk", "Normal", "", "TRUE"]
    required = [1, 2]
    add_header(ws, headers, required, hints)
    add_validation_list(ws, 1, 3, MAX_ROWS, named['ServiceCategories'])
    add_validation_list(ws, 2, 3, MAX_ROWS, named['Departments'])
    add_validation_list(ws, 6, 3, MAX_ROWS, named['TaskPriority'])
    add_validation_list(ws, 8, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 3, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 4, 3, MAX_ROWS, named['YesNo'])
    ws.cell(1, 1).comment = Comment("Category of service a guest requests (e.g., Dining).", "Hostr")
    ws.cell(1, 2).comment = Comment("Which department should handle the request?", "Hostr")
    set_col_widths(ws, [22, 20, 12, 18, 18, 12, 28, 10])
    return ws


def sheet_request_items(wb, named):
    ws = wb.create_sheet("Request Items")
    headers = [
        "Name*", "Category", "Department", "Purpose", "Description", "IsAvailable", "RequiresQuantity", "DefaultQuantityLimit", "RequiresRoomDelivery",
        "StockCount", "InServiceCount", "AutoDecrementOnTask", "LowStockThreshold", "LlmVisibleName", "NotesForStaff", "ServiceHours (JSON)", "SlaMinutes", "EstimatedTime", "IsUrgent", "DisplayOrder"
    ]
    hints = [
        "Extra Towels", "Amenities", "Housekeeping", "", "", "TRUE", "TRUE", "10", "TRUE",
        "0", "0", "FALSE", "5", "Towels", "", "{...}", "30", "", "FALSE", "0"
    ]
    required = [1]
    add_header(ws, headers, required, hints)
    add_validation_list(ws, 3, 3, MAX_ROWS, named['Departments'])
    add_validation_list(ws, 6, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 7, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 9, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 12, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 19, 3, MAX_ROWS, named['YesNo'])
    ws.cell(1, 1).comment = Comment("What the guest will ask for (e.g., Extra Towels).", "Hostr")
    ws.cell(1, 3).comment = Comment("Which department fulfills this item.", "Hostr")
    ws.cell(1, 17).comment = Comment("Target SLA in minutes.", "Hostr")
    set_col_widths(ws, [20, 16, 16, 16, 26, 10, 14, 16, 16, 12, 14, 18, 18, 18, 20, 22, 12, 14, 10, 12])
    return ws


def sheet_concierge_providers(wb, named):
    ws = wb.create_sheet("Concierge & Providers")
    headers = [
        "ServiceCategoryName", "ServiceName", "ServiceDescription", "ServiceLlmVisibleName", "ServiceIsActive", "RequiresAdvanceNotice", "AdvanceNoticeText", "ResponseTemplate",
        "ProviderName", "ProviderDescription", "ProviderPhone", "ProviderEmail", "ProviderAddress", "ProviderWebsite", "IsRecommended", "ProviderIsActive", "DisplayOrder", "Notes"
    ]
    hints = [
        "Transportation", "Airport Shuttle", "Shuttle service to/from airport", "Airport Shuttle", "TRUE", "FALSE", "", "",
        "Shuttle Co.", "Reliable shuttle", "+27...", "info@shuttle.co", "123 Road, City", "https://...", "TRUE", "TRUE", "0", ""
    ]
    add_header(ws, headers, [], hints)
    add_validation_list(ws, 1, 3, MAX_ROWS, named['ServiceCategories'])
    add_validation_list(ws, 5, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 6, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 15, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 16, 3, MAX_ROWS, named['YesNo'])
    set_col_widths(ws, [22, 22, 30, 20, 10, 14, 24, 24, 20, 26, 16, 22, 26, 24, 12, 12, 12, 24])
    return ws


def sheet_business_info(wb, named):
    ws = wb.create_sheet("Business Info")
    headers = ["Category", "Title", "Content", "Tags", "IsActive", "DisplayOrder"]
    hints = ["hours/location/...", "Breakfast Hours", "Breakfast served 6am-10am", "breakfast,morning", "TRUE", "0"]
    add_header(ws, headers, [], hints)
    add_validation_list(ws, 1, 3, MAX_ROWS, named['InfoCategories'])
    add_validation_list(ws, 5, 3, MAX_ROWS, named['YesNo'])
    set_col_widths(ws, [16, 24, 60, 20, 10, 12])
    return ws


def sheet_faq_knowledge(wb, named):
    ws = wb.create_sheet("FAQ Knowledge")
    headers = ["Question", "Answer", "Category", "Keywords", "IsTimeRelevant", "RelevantHourStart", "RelevantHourEnd", "IsActive", "Priority"]
    hints = ["Do you have parking?", "Yes, underground parking available.", "general", "parking,car", "FALSE", "", "", "TRUE", "0"]
    add_header(ws, headers, [], hints)
    add_validation_list(ws, 3, 3, MAX_ROWS, named['FAQCategories'])
    add_validation_list(ws, 5, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 8, 3, MAX_ROWS, named['YesNo'])
    set_col_widths(ws, [32, 60, 16, 24, 12, 16, 16, 10, 10])
    return ws


def sheet_menu(wb, named):
    ws = wb.create_sheet("Menu")
    headers = [
        "CategoryName", "CategoryDescription", "CategoryMealType", "CategoryDisplayOrder", "CategoryIsActive",
        "ItemName", "ItemDescription", "PriceCents", "Currency", "Allergens", "ItemMealType", "IsVegetarian", "IsVegan", "IsGlutenFree", "IsSpicy", "IsAvailable", "IsSpecial", "Tags", "ImageUrl", "ItemDisplayOrder",
        "SpecialTitle", "SpecialDescription", "SpecialPriceCents", "SpecialType", "SpecialDayOfWeek", "ValidFrom", "ValidTo", "SpecialMealType", "SpecialIsActive"
    ]
    hints = [
        "Breakfast", "", "all", "0", "TRUE",
        "Pancakes", "Fluffy pancakes", "8000", "ZAR", "gluten,eggs", "all", "FALSE", "FALSE", "FALSE", "FALSE", "TRUE", "FALSE", "sweet,classic", "https://...", "0",
        "Daily Pancake Special", "Buy 1 get 1", "7000", "daily", "", "", "", "all", "TRUE"
    ]
    add_header(ws, headers, [], hints)
    add_validation_list(ws, 3, 3, MAX_ROWS, named['MealTypes'])
    add_validation_list(ws, 11, 3, MAX_ROWS, named['MealTypes'])
    add_validation_list(ws, 9, 3, MAX_ROWS, named['Currencies'])
    add_validation_list(ws, 5, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 12, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 13, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 14, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 15, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 16, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 17, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 29, 3, MAX_ROWS, named['YesNo'])
    set_col_widths(ws, [18, 26, 12, 10, 10, 22, 30, 12, 10, 18, 12, 10, 10, 12, 10, 10, 10, 20, 24, 12, 24, 30, 12, 12, 12, 12, 12, 12, 10])
    return ws


def sheet_templates(wb, named):
    ws = wb.create_sheet("Templates")
    headers = ["WaTemplateName", "WaCategory", "WaLanguage", "WaBody", "WaStatus", "QuickReplyTitle", "QuickReplyBody"]
    hints = ["greeting_template", "AUTHENTICATION", "en", "Hello ...", "Pending", "Welcome", "Welcome to our hotel..."]
    add_header(ws, headers, [], hints)
    set_col_widths(ws, [24, 18, 10, 50, 14, 20, 50])
    return ws


def sheet_emergency(wb, named):
    ws = wb.create_sheet("Emergency")
    headers = [
        "TypeName", "TypeDescription", "DetectionKeywords", "SeverityLevel", "AutoEscalate", "RequiresEvacuation", "ContactEmergencyServices", "IsActive",
        "ContactName", "ContactType", "PhoneNumber", "Email", "Address", "Notes", "IsPrimary", "ContactIsActive", "IsPublic",
        "ProtocolEmergencyType", "ProtocolTitle", "ProcedureSteps", "TriggerCondition", "NotifyGuests", "NotifyStaff", "NotifyEmergencyServices", "GuestMessage", "StaffMessage", "EmergencyContacts (comma)", "ProtocolIsActive", "ExecutionOrder"
    ]
    hints = [
        "Medical", "Medical emergencies", "injury,help,ambulance", "High", "TRUE", "FALSE", "FALSE", "TRUE",
        "Fire Dept.", "Fire Department", "+27...", "", "", "", "TRUE", "TRUE", "FALSE",
        "Fire", "Evacuation Procedure", "Steps...", "IMMEDIATE", "FALSE", "TRUE", "FALSE", "", "", "Chief,Security", "TRUE", "1"
    ]
    add_header(ws, headers, [], hints)
    add_validation_list(ws, 4, 3, MAX_ROWS, named['SeverityLevels'])
    add_validation_list(ws, 5, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 6, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 7, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 8, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 15, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 16, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 17, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 21, 3, MAX_ROWS, named['TriggerCondition'])
    add_validation_list(ws, 22, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 23, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 24, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 27, 3, MAX_ROWS, named['YesNo'])
    set_col_widths(ws, [16, 26, 24, 12, 10, 10, 10, 10, 16, 18, 16, 22, 26, 22, 10, 12, 10, 20, 26, 30, 16, 10, 10, 10, 22, 22, 24, 10, 12])
    return ws


def sheet_property_directions(wb, named):
    ws = wb.create_sheet("Property Directions")
    headers = ["FacilityName*", "Category*", "Directions*", "LocationDescription", "Floor", "Wing", "Landmarks", "Hours", "IsActive", "Priority", "ImageUrl", "AdditionalInfo"]
    hints = ["Pool", "Pool/Dining/Spa/Fitness", "From lobby, turn left...", "Ground floor near lobby", "Ground Floor", "East Wing", "Next to elevator", "06:00-22:00", "TRUE", "0", "https://...", ""]
    required = [1, 2, 3]
    add_header(ws, headers, required, hints)
    add_validation_list(ws, 9, 3, MAX_ROWS, named['YesNo'])
    set_col_widths(ws, [18, 16, 40, 26, 16, 16, 26, 16, 10, 10, 26, 26])
    return ws


def sheet_intent_settings(wb, named):
    ws = wb.create_sheet("Intent Settings")
    headers = [
        "IntentName*", "IsEnabled", "EnableUpselling", "CustomResponse", "UpsellStrategy", "Priority", "RequiresStaffApproval", "NotifyStaff", "AssignedDepartment", "TaskPriority", "AutoResolve", "AutoResolveDelayMinutes", "AdditionalConfig (JSON)"
    ]
    hints = ["CHECK_IN_OUT", "TRUE", "TRUE", "", "{}", "0", "FALSE", "TRUE", "FrontDesk", "Medium", "FALSE", "", "{}"]
    required = [1]
    add_header(ws, headers, required, hints)
    add_validation_list(ws, 1, 3, MAX_ROWS, named['IntentNames'])
    add_validation_list(ws, 2, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 3, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 7, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 8, 3, MAX_ROWS, named['YesNo'])
    add_validation_list(ws, 9, 3, MAX_ROWS, named['Departments'])
    add_validation_list(ws, 10, 3, MAX_ROWS, named['TaskPriority'])
    add_validation_list(ws, 11, 3, MAX_ROWS, named['YesNo'])
    set_col_widths(ws, [16, 10, 14, 40, 26, 10, 16, 12, 18, 12, 10, 20, 26])
    return ws


def sheet_welcome_messages(wb, named):
    ws = wb.create_sheet("Welcome Messages")
    headers = ["MessageType*", "Template*", "IsActive", "DisplayOrder"]
    hints = ["greeting/welcome/assistance", "Hi {{guest_name}}...", "TRUE", "0"]
    required = [1, 2]
    add_header(ws, headers, required, hints)
    return ws

def prefill_from_db(conn, wb, tenant_id, debug=False):
    if not conn:
        return
    # Core Tenant
    rows = try_queries(conn, [
        'SELECT "Name", "Slug", "Timezone", "Plan", "ThemePrimary", "Status", "RetentionDays" FROM "Tenants" WHERE "Id"=%s',
        'SELECT name as "Name", slug as "Slug", timezone as "Timezone", plan as "Plan", theme_primary as "ThemePrimary", status as "Status", retention_days as "RetentionDays" FROM tenants WHERE id=%s'
    ], [tenant_id], debug, label="Tenants")
    if rows:
        r = rows[0]
        ws = wb["Core Tenant"]
        ws.cell(3,1,r.get("Name"))
        ws.cell(3,2,r.get("Slug"))
        ws.cell(3,3,r.get("Timezone"))
        ws.cell(3,4,r.get("Plan"))
        ws.cell(3,5,r.get("ThemePrimary"))
        ws.cell(3,6,r.get("Status"))
        ws.cell(3,7,r.get("RetentionDays"))

    # WhatsApp (first number)
    rows = try_queries(conn, [
        'SELECT "WabaId", "PhoneNumberId", "PageAccessToken", "Status" FROM "WhatsAppNumbers" WHERE "TenantId"=%s ORDER BY "Id" LIMIT 1',
        'SELECT waba_id as "WabaId", phone_number_id as "PhoneNumberId", page_access_token as "PageAccessToken", status as "Status" FROM whatsapp_numbers WHERE tenant_id=%s ORDER BY id LIMIT 1'
    ], [tenant_id], debug, label="WhatsAppNumbers")
    if rows:
        r = rows[0]
        ws = wb["WhatsApp (WABA)"]
        ws.cell(3,1,r.get("WabaId"))
        ws.cell(3,2,r.get("PhoneNumberId"))
        ws.cell(3,3,r.get("PageAccessToken"))
        ws.cell(3,5,r.get("Status"))

    # Hotel Info (single row)
    rows = try_queries(conn, [
        'SELECT * FROM "HotelInfos" WHERE "TenantId"=%s ORDER BY "Id" LIMIT 1',
        'SELECT * FROM hotel_infos WHERE tenant_id=%s ORDER BY id LIMIT 1'
    ], [tenant_id], debug, label="HotelInfos")
    if rows:
        r = rows[0]
        ws = wb["Hotel Info"]
        ws.cell(3,1,r.get("Description"))
        ws.cell(3,2,r.get("Category"))
        ws.cell(3,3,r.get("LogoUrl"))
        ws.cell(3,4,r.get("Phone"))
        ws.cell(3,5,r.get("Email"))
        ws.cell(3,6,r.get("Website"))
        ws.cell(3,7,r.get("Street"))
        ws.cell(3,8,r.get("City"))
        ws.cell(3,9,r.get("State"))
        ws.cell(3,10,r.get("PostalCode"))
        ws.cell(3,11,r.get("Country"))
        ws.cell(3,12,r.get("CheckInTime"))
        ws.cell(3,13,r.get("CheckOutTime"))
        ws.cell(3,14,r.get("NumberOfRooms"))
        ws.cell(3,15,r.get("NumberOfFloors"))
        ws.cell(3,16,r.get("EstablishedYear"))
        ws.cell(3,17,r.get("SupportedLanguages"))
        ws.cell(3,18,r.get("DefaultLanguage"))
        ws.cell(3,19,r.get("Features"))
        ws.cell(3,20,r.get("FacebookUrl"))
        ws.cell(3,21,r.get("TwitterUrl"))
        ws.cell(3,22,r.get("InstagramUrl"))
        ws.cell(3,23,r.get("LinkedInUrl"))
        ws.cell(3,24,r.get("CancellationPolicy"))
        ws.cell(3,25,r.get("PetPolicy"))
        ws.cell(3,26,r.get("SmokingPolicy"))
        ws.cell(3,27,r.get("ChildPolicy"))
        ws.cell(3,28,r.get("AllowOnlineBooking"))
        ws.cell(3,29,r.get("RequirePhoneVerification"))
        ws.cell(3,30,r.get("EnableNotifications"))
        ws.cell(3,31,r.get("EnableChatbot"))
        ws.cell(3,32,r.get("Currency"))

    # Tenant Settings
    rows = try_queries(conn, [
        'SELECT * FROM "TenantSettings" WHERE "TenantId"=%s ORDER BY "Id" LIMIT 1',
        'SELECT * FROM tenant_settings WHERE tenant_id=%s ORDER BY id LIMIT 1'
    ], [tenant_id], debug, label="TenantSettings")
    if rows:
        r = rows[0]
        ws = wb["Tenant Settings"]
        def ts_to_str(val):
            try:
                return str(val) if val is None else f"{val}"
            except Exception:
                return None
        ws.cell(3,1, ts_to_str(r.get("BusinessHoursStart")))
        ws.cell(3,2, ts_to_str(r.get("BusinessHoursEnd")))
        ws.cell(3,3, ts_to_str(r.get("StandardCheckInTime")))
        ws.cell(3,4, ts_to_str(r.get("StandardCheckOutTime")))
        ws.cell(3,5, r.get("LateCheckoutFeePerHour"))
        ws.cell(3,6, r.get("EarlyCheckInFeePerHour"))
        ws.cell(3,7, r.get("DefaultCurrency"))
        ws.cell(3,8, r.get("Timezone"))

    # Departments
    rows = try_queries(conn, [
        'SELECT * FROM "TenantDepartments" WHERE "TenantId"=%s ORDER BY "Priority", "DepartmentName"',
        'SELECT * FROM tenant_departments WHERE tenant_id=%s ORDER BY priority, department_name'
    ], [tenant_id], debug, label="TenantDepartments")
    if rows:
        ws = wb["Departments"]
        r0 = 3
        for r in rows:
            ws.cell(r0,1,r.get("DepartmentName")); ws.cell(r0,2,r.get("Description")); ws.cell(r0,3,r.get("IsActive"))
            ws.cell(r0,4,r.get("Priority")); ws.cell(r0,5,r.get("ContactInfo")); ws.cell(r0,6,r.get("WorkingHours")); ws.cell(r0,7,r.get("MaxConcurrentTasks"))
            r0 += 1

    # Request Items
    rows = try_queries(conn, [
        'SELECT * FROM "RequestItems" WHERE "TenantId"=%s ORDER BY "DisplayOrder", "Name" LIMIT 100',
        'SELECT * FROM request_items WHERE tenant_id=%s ORDER BY display_order, name LIMIT 100'
    ], [tenant_id], debug, label="RequestItems")
    if rows:
        ws = wb["Request Items"]
        r0 = 3
        for r in rows:
            vals = [
                r.get("Name"), r.get("Category"), r.get("Department"), r.get("Purpose"), r.get("Description"), r.get("IsAvailable"), r.get("RequiresQuantity"), r.get("DefaultQuantityLimit"), r.get("RequiresRoomDelivery"),
                r.get("StockCount"), r.get("InServiceCount"), r.get("AutoDecrementOnTask"), r.get("LowStockThreshold"), r.get("LlmVisibleName"), r.get("NotesForStaff"), r.get("ServiceHours"), r.get("SlaMinutes"), r.get("EstimatedTime"), r.get("IsUrgent"), r.get("DisplayOrder")
            ]
            for i,v in enumerate(vals, start=1): ws.cell(r0,i,v)
            r0 += 1

    # Menu (Categories + Items + Specials) using provided queries
    cat_rows = safe_query(conn, 'SELECT "Id", "TenantId", "Name", "Description", "MealType", "DisplayOrder", "IsActive", "UpdatedAt" FROM public."MenuCategories" WHERE "TenantId"=%s ORDER BY "DisplayOrder", "Name"', [tenant_id])
    item_rows = safe_query(conn, 'SELECT "Id", "TenantId", "MenuCategoryId", "Name", "Description", "PriceCents", "Currency", "Allergens", "MealType", "IsVegetarian", "IsVegan", "IsGlutenFree", "IsSpicy", "IsAvailable", "IsSpecial", "Tags", "UpdatedAt" FROM public."MenuItems" WHERE "TenantId"=%s ORDER BY "MenuCategoryId", "Name"', [tenant_id])
    special_rows = safe_query(conn, 'SELECT "Id", "TenantId", "MenuItemId", "Title", "Description", "SpecialPriceCents", "SpecialType", "DayOfWeek", "ValidFrom", "ValidTo", "MealType", "IsActive", "UpdatedAt" FROM public."MenuSpecials" WHERE "TenantId"=%s ORDER BY "MenuItemId", "Title"', [tenant_id])

    if cat_rows or item_rows or special_rows:
        if debug:
            print(f"[prefill] Menu -> {len(cat_rows)} categories, {len(item_rows)} items, {len(special_rows)} specials")
        ws = wb["Menu"]
        # Index categories and specials
        cat_by_id = {c.get("Id"): c for c in cat_rows}
        specials_by_item = {}
        for s in special_rows:
            specials_by_item.setdefault(s.get("MenuItemId"), []).append(s)

        r0 = 3
        # For each item, write one row, and extra rows for specials (if any)
        for it in item_rows:
            cid = it.get("MenuCategoryId")
            c = cat_by_id.get(cid, {})

            # Category fields
            cat_name = c.get("Name")
            cat_desc = c.get("Description")
            cat_meal = c.get("MealType")
            cat_order = c.get("DisplayOrder")
            cat_active = c.get("IsActive")

            # Item fields
            name = it.get("Name")
            desc = it.get("Description")
            price_cents = it.get("PriceCents")
            currency = it.get("Currency")
            allergens = it.get("Allergens")
            item_meal = it.get("MealType")
            is_veg = it.get("IsVegetarian")
            is_vegan = it.get("IsVegan")
            is_gf = it.get("IsGlutenFree")
            is_spicy = it.get("IsSpicy")
            is_available = it.get("IsAvailable")
            is_special = it.get("IsSpecial")
            tags = it.get("Tags")
            image_url = None
            item_order = None

            # Base row without specials
            base_vals = [
                cat_name, cat_desc, cat_meal, cat_order, cat_active,
                name, desc, price_cents, currency, allergens, item_meal, is_veg, is_vegan, is_gf, is_spicy, is_available, is_special, tags, image_url, item_order
            ]

            item_specials = specials_by_item.get(it.get("Id"), [])
            if not item_specials:
                # no specials
                vals = base_vals + [None, None, None, None, None, None, None, None, None]
                for i,v in enumerate(vals, start=1): ws.cell(r0,i,v)
                r0 += 1
            else:
                for s in item_specials:
                    special_title = s.get("Title")
                    special_desc = s.get("Description")
                    special_price = s.get("SpecialPriceCents")
                    special_type = s.get("SpecialType")
                    special_dow = s.get("DayOfWeek")
                    valid_from = s.get("ValidFrom")
                    valid_to = s.get("ValidTo")
                    special_meal = s.get("MealType")
                    special_active = s.get("IsActive")
                    vals = base_vals + [special_title, special_desc, special_price, special_type, special_dow, valid_from, valid_to, special_meal, special_active]
                    for i,v in enumerate(vals, start=1): ws.cell(r0,i,v)
                    r0 += 1

    # Business Info
    rows = try_queries(conn, [
        'SELECT * FROM "BusinessInfo" WHERE "TenantId"=%s ORDER BY "DisplayOrder", "Title" LIMIT 100',
        'SELECT * FROM business_info WHERE tenant_id=%s ORDER BY display_order, title LIMIT 100'
    ], [tenant_id], debug, label="BusinessInfo")
    if rows:
        ws = wb["Business Info"]
        r0 = 3
        for r in rows:
            ws.cell(r0,1,r.get("Category")); ws.cell(r0,2,r.get("Title")); ws.cell(r0,3,r.get("Content")); ws.cell(r0,4, ",".join(r.get("Tags") or [])); ws.cell(r0,5,r.get("IsActive")); ws.cell(r0,6,r.get("DisplayOrder"))
            r0 += 1

    # FAQ Knowledge (InformationItems)
    rows = try_queries(conn, [
        'SELECT * FROM "InformationItems" WHERE "TenantId"=%s ORDER BY "Priority" DESC, "UpdatedAt" DESC LIMIT 200',
        'SELECT * FROM information_items WHERE tenant_id=%s ORDER BY priority DESC, updated_at DESC LIMIT 200'
    ], [tenant_id], debug, label="InformationItems")
    if rows:
        ws = wb["FAQ Knowledge"]
        r0 = 3
        for r in rows:
            ws.cell(r0,1,r.get("Question")); ws.cell(r0,2,r.get("Answer")); ws.cell(r0,3,r.get("Category")); ws.cell(r0,4, ",".join(r.get("Keywords") or []))
            ws.cell(r0,5,r.get("IsTimeRelevant")); ws.cell(r0,6,r.get("RelevantHourStart")); ws.cell(r0,7,r.get("RelevantHourEnd")); ws.cell(r0,8,r.get("IsActive")); ws.cell(r0,9,r.get("Priority"))
            r0 += 1

    # Intent Settings
    rows = try_queries(conn, [
        'SELECT * FROM "IntentSettings" WHERE "TenantId"=%s ORDER BY "Priority" DESC, "IntentName"',
        'SELECT * FROM intent_settings WHERE tenant_id=%s ORDER BY priority DESC, intent_name'
    ], [tenant_id], debug, label="IntentSettings")
    if rows:
        ws = wb["Intent Settings"]
        r0 = 3
        for r in rows:
            vals = [r.get("IntentName"), r.get("IsEnabled"), r.get("EnableUpselling"), r.get("CustomResponse"), r.get("UpsellStrategy"), r.get("Priority"), r.get("RequiresStaffApproval"), r.get("NotifyStaff"), r.get("AssignedDepartment"), r.get("TaskPriority"), r.get("AutoResolve"), r.get("AutoResolveDelayMinutes"), r.get("AdditionalConfig")]
            for i,v in enumerate(vals, start=1): ws.cell(r0,i,v)
            r0 += 1

    # Property Directions
    rows = try_queries(conn, [
        'SELECT * FROM "PropertyDirections" WHERE "TenantId"=%s ORDER BY "Priority", "FacilityName"',
        'SELECT * FROM property_directions WHERE tenant_id=%s ORDER BY priority, facility_name'
    ], [tenant_id], debug, label="PropertyDirections")
    if rows:
        ws = wb["Property Directions"]
        r0 = 3
        for r in rows:
            vals = [r.get("FacilityName"), r.get("Category"), r.get("Directions"), r.get("LocationDescription"), r.get("Floor"), r.get("Wing"), r.get("Landmarks"), r.get("Hours"), r.get("IsActive"), r.get("Priority"), r.get("ImageUrl"), r.get("AdditionalInfo")]
            for i,v in enumerate(vals, start=1): ws.cell(r0,i,v)
            r0 += 1

    # Welcome Messages
    rows = try_queries(conn, [
        'SELECT * FROM "WelcomeMessages" WHERE "TenantId"=%s ORDER BY "DisplayOrder"',
        'SELECT * FROM welcome_messages WHERE tenant_id=%s ORDER BY display_order'
    ], [tenant_id], debug, label="WelcomeMessages")
    if rows:
        ws = wb["Welcome Messages"]
        r0 = 3
        for r in rows:
            ws.cell(r0,1,r.get("MessageType")); ws.cell(r0,2,r.get("Template")); ws.cell(r0,3,r.get("IsActive")); ws.cell(r0,4,r.get("DisplayOrder"))
            r0 += 1


def main(out_path, conn_str=None, tenant_id=1, debug=False):
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    named = build_lists_sheet(wb)

    build_start_here(wb)
    sheet_core_tenant(wb, named)
    sheet_waba(wb, named)
    sheet_owner_staff(wb, named)
    sheet_hotel_info(wb, named)
    sheet_tenant_settings(wb, named)
    sheet_departments(wb, named)
    sheet_service_mapping(wb, named)
    sheet_request_items(wb, named)
    sheet_concierge_providers(wb, named)
    sheet_business_info(wb, named)
    sheet_faq_knowledge(wb, named)
    sheet_menu(wb, named)
    sheet_templates(wb, named)
    sheet_emergency(wb, named)
    sheet_property_directions(wb, named)
    sheet_intent_settings(wb, named)
    sheet_welcome_messages(wb, named)

    # Optional DB prefill
    raw_conn = conn_str or os.environ.get("STAYBOT_CONN", "")
    if debug:
        print("[prefill] Using connection string from", "--conn" if conn_str else "STAYBOT_CONN env")
    conn = get_connection(raw_conn, debug=debug)
    try:
        if debug and not conn:
            print("[prefill] No DB connection available; skipping prefill")
        prefill_from_db(conn, wb, tenant_id, debug=debug)
    finally:
        try:
            if conn:
                conn.close()
        except Exception:
            pass

    wb.save(out_path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Tenant Onboarding Excel")
    parser.add_argument("--conn", help="PostgreSQL connection string (or set STAYBOT_CONN)", default=None)
    parser.add_argument("--tenant-id", type=int, help="Tenant ID to prefill", default=1)
    parser.add_argument("--out", help="Output xlsx path", default=None)
    parser.add_argument("--debug", action="store_true", help="Enable debug logging for prefill queries")
    args = parser.parse_args()

    out_file = args.out or os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "Tenant_Onboarding_Questionnaire.xlsx"))
    os.makedirs(os.path.dirname(out_file), exist_ok=True)
    main(out_file, conn_str=args.conn, tenant_id=args.tenant_id, debug=args.debug)
    print(f"Generated: {out_file}")
